#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel工具软件 - 局域网Excel便捷工具
主要功能：Excel列对比、数据去重、格式调整等
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import threading
from datetime import datetime
import json

class ExcelTool:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel工具软件 - 局域网Excel便捷工具")
        self.root.geometry("1000x700")
        self.root.resizable(True, True)
        
        # 设置图标和样式
        self.setup_styles()
        
        # 初始化变量
        self.file_a_path = tk.StringVar()
        self.file_b_path = tk.StringVar()
        self.column_a = tk.StringVar()
        self.column_b = tk.StringVar()
        self.progress_var = tk.DoubleVar()
        
        # 存储数据
        self.df_a = None
        self.df_b = None
        self.columns_a = []
        self.columns_b = []
        
        # 创建界面
        self.create_widgets()
        
    def setup_styles(self):
        """设置界面样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置样式
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'))
        style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Success.TLabel', foreground='green')
        style.configure('Error.TLabel', foreground='red')
        
    def create_widgets(self):
        """创建界面组件"""
        # 主标题
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill='x', padx=10, pady=10)
        
        title_label = ttk.Label(title_frame, text="Excel工具软件", style='Title.TLabel')
        title_label.pack()
        
        subtitle_label = ttk.Label(title_frame, text="局域网Excel便捷工具 - 无需外网连接", 
                                 font=('Arial', 10))
        subtitle_label.pack()
        
        # 创建notebook用于多标签页
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # 列对比功能页面
        self.create_column_compare_tab()
        
        # 数据去重功能页面
        self.create_deduplication_tab()
        
        # 格式调整功能页面
        self.create_format_tab()
        
        # 状态栏
        self.create_status_bar()
        
    def create_column_compare_tab(self):
        """创建列对比功能标签页"""
        compare_frame = ttk.Frame(self.notebook)
        self.notebook.add(compare_frame, text="列对比功能")
        
        # 文件上传区域
        file_frame = ttk.LabelFrame(compare_frame, text="文件上传", padding=10)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        # 文件A
        file_a_frame = ttk.Frame(file_frame)
        file_a_frame.pack(fill='x', pady=5)
        
        ttk.Label(file_a_frame, text="文件A:", width=10).pack(side='left')
        ttk.Entry(file_a_frame, textvariable=self.file_a_path, width=50).pack(side='left', padx=5)
        ttk.Button(file_a_frame, text="选择文件", 
                  command=lambda: self.select_file('a')).pack(side='left')
        
        # 文件B
        file_b_frame = ttk.Frame(file_frame)
        file_b_frame.pack(fill='x', pady=5)
        
        ttk.Label(file_b_frame, text="文件B:", width=10).pack(side='left')
        ttk.Entry(file_b_frame, textvariable=self.file_b_path, width=50).pack(side='left', padx=5)
        ttk.Button(file_b_frame, text="选择文件", 
                  command=lambda: self.select_file('b')).pack(side='left')
        
        # 列选择区域
        column_frame = ttk.LabelFrame(compare_frame, text="列选择", padding=10)
        column_frame.pack(fill='x', padx=10, pady=5)
        
        # 列A选择
        col_a_frame = ttk.Frame(column_frame)
        col_a_frame.pack(fill='x', pady=5)
        
        ttk.Label(col_a_frame, text="文件A列:", width=10).pack(side='left')
        self.combo_a = ttk.Combobox(col_a_frame, textvariable=self.column_a, width=30)
        self.combo_a.pack(side='left', padx=5)
        
        # 列B选择
        col_b_frame = ttk.Frame(column_frame)
        col_b_frame.pack(fill='x', pady=5)
        
        ttk.Label(col_b_frame, text="文件B列:", width=10).pack(side='left')
        self.combo_b = ttk.Combobox(col_b_frame, textvariable=self.column_b, width=30)
        self.combo_b.pack(side='left', padx=5)
        
        # 操作按钮
        button_frame = ttk.Frame(compare_frame)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(button_frame, text="加载文件", 
                  command=self.load_files).pack(side='left', padx=5)
        ttk.Button(button_frame, text="开始对比", 
                  command=self.start_compare).pack(side='left', padx=5)
        ttk.Button(button_frame, text="保存结果", 
                  command=self.save_result).pack(side='left', padx=5)
        
        # 进度条
        progress_frame = ttk.Frame(compare_frame)
        progress_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(progress_frame, text="处理进度:").pack(side='left')
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                          maximum=100)
        self.progress_bar.pack(side='left', fill='x', expand=True, padx=5)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(compare_frame, text="对比结果", padding=10)
        result_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.result_text = scrolledtext.ScrolledText(result_frame, height=15)
        self.result_text.pack(fill='both', expand=True)
        
    def create_deduplication_tab(self):
        """创建数据去重功能标签页"""
        dedup_frame = ttk.Frame(self.notebook)
        self.notebook.add(dedup_frame, text="数据去重")
        
        # 文件上传区域
        file_frame = ttk.LabelFrame(dedup_frame, text="文件上传", padding=10)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        self.dedup_file_path = tk.StringVar()
        
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill='x', pady=5)
        
        ttk.Label(file_select_frame, text="文件:", width=10).pack(side='left')
        ttk.Entry(file_select_frame, textvariable=self.dedup_file_path, width=50).pack(side='left', padx=5)
        ttk.Button(file_select_frame, text="选择文件", 
                  command=self.select_dedup_file).pack(side='left')
        
        # 列选择区域
        dedup_column_frame = ttk.LabelFrame(dedup_frame, text="去重设置", padding=10)
        dedup_column_frame.pack(fill='x', padx=10, pady=5)
        
        # 去重列选择
        dedup_col_frame = ttk.Frame(dedup_column_frame)
        dedup_col_frame.pack(fill='x', pady=5)
        
        ttk.Label(dedup_col_frame, text="去重列:", width=10).pack(side='left')
        self.dedup_column = tk.StringVar()
        self.dedup_combo = ttk.Combobox(dedup_col_frame, textvariable=self.dedup_column, width=30)
        self.dedup_combo.pack(side='left', padx=5)
        
        # 保留策略
        strategy_frame = ttk.Frame(dedup_column_frame)
        strategy_frame.pack(fill='x', pady=5)
        
        ttk.Label(strategy_frame, text="保留策略:", width=10).pack(side='left')
        self.keep_strategy = tk.StringVar(value="first")
        ttk.Radiobutton(strategy_frame, text="保留第一条", variable=self.keep_strategy, 
                       value="first").pack(side='left', padx=5)
        ttk.Radiobutton(strategy_frame, text="保留最后一条", variable=self.keep_strategy, 
                       value="last").pack(side='left', padx=5)
        
        # 操作按钮
        dedup_button_frame = ttk.Frame(dedup_frame)
        dedup_button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(dedup_button_frame, text="加载文件", 
                  command=self.load_dedup_file).pack(side='left', padx=5)
        ttk.Button(dedup_button_frame, text="开始去重", 
                  command=self.start_deduplication).pack(side='left', padx=5)
        ttk.Button(dedup_button_frame, text="保存结果", 
                  command=self.save_dedup_result).pack(side='left', padx=5)
        
        # 结果显示区域
        dedup_result_frame = ttk.LabelFrame(dedup_frame, text="去重结果", padding=10)
        dedup_result_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.dedup_result_text = scrolledtext.ScrolledText(dedup_result_frame, height=15)
        self.dedup_result_text.pack(fill='both', expand=True)
        
    def create_format_tab(self):
        """创建格式调整功能标签页"""
        format_frame = ttk.Frame(self.notebook)
        self.notebook.add(format_frame, text="格式调整")
        
        # 文件上传区域
        format_file_frame = ttk.LabelFrame(format_frame, text="文件上传", padding=10)
        format_file_frame.pack(fill='x', padx=10, pady=5)
        
        self.format_file_path = tk.StringVar()
        
        format_file_select_frame = ttk.Frame(format_file_frame)
        format_file_select_frame.pack(fill='x', pady=5)
        
        ttk.Label(format_file_select_frame, text="文件:", width=10).pack(side='left')
        ttk.Entry(format_file_select_frame, textvariable=self.format_file_path, width=50).pack(side='left', padx=5)
        ttk.Button(format_file_select_frame, text="选择文件", 
                  command=self.select_format_file).pack(side='left')
        
        # 格式设置区域
        format_settings_frame = ttk.LabelFrame(format_frame, text="格式设置", padding=10)
        format_settings_frame.pack(fill='x', padx=10, pady=5)
        
        # 列选择
        format_col_frame = ttk.Frame(format_settings_frame)
        format_col_frame.pack(fill='x', pady=5)
        
        ttk.Label(format_col_frame, text="目标列:", width=10).pack(side='left')
        self.format_column = tk.StringVar()
        self.format_combo = ttk.Combobox(format_col_frame, textvariable=self.format_column, width=30)
        self.format_combo.pack(side='left', padx=5)
        
        # 格式类型选择
        format_type_frame = ttk.Frame(format_settings_frame)
        format_type_frame.pack(fill='x', pady=5)
        
        ttk.Label(format_type_frame, text="格式类型:", width=10).pack(side='left')
        self.format_type = tk.StringVar(value="date")
        ttk.Radiobutton(format_type_frame, text="日期格式", variable=self.format_type, 
                       value="date").pack(side='left', padx=5)
        ttk.Radiobutton(format_type_frame, text="数字格式", variable=self.format_type, 
                       value="number").pack(side='left', padx=5)
        ttk.Radiobutton(format_type_frame, text="文本格式", variable=self.format_type, 
                       value="text").pack(side='left', padx=5)
        
        # 具体格式设置
        self.format_detail_frame = ttk.Frame(format_settings_frame)
        self.format_detail_frame.pack(fill='x', pady=5)
        
        # 操作按钮
        format_button_frame = ttk.Frame(format_frame)
        format_button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(format_button_frame, text="加载文件", 
                  command=self.load_format_file).pack(side='left', padx=5)
        ttk.Button(format_button_frame, text="应用格式", 
                  command=self.apply_format).pack(side='left', padx=5)
        ttk.Button(format_button_frame, text="保存结果", 
                  command=self.save_format_result).pack(side='left', padx=5)
        
        # 结果显示区域
        format_result_frame = ttk.LabelFrame(format_frame, text="格式调整结果", padding=10)
        format_result_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.format_result_text = scrolledtext.ScrolledText(format_result_frame, height=15)
        self.format_result_text.pack(fill='both', expand=True)
        
    def create_status_bar(self):
        """创建状态栏"""
        self.status_bar = ttk.Label(self.root, text="就绪", relief='sunken')
        self.status_bar.pack(side='bottom', fill='x')
        
    def select_file(self, file_type):
        """选择文件"""
        file_path = filedialog.askopenfilename(
            title=f"选择文件{file_type.upper()}",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if file_path:
            if file_type == 'a':
                self.file_a_path.set(file_path)
            else:
                self.file_b_path.set(file_path)
                
    def load_files(self):
        """加载Excel文件"""
        try:
            self.status_bar.config(text="正在加载文件...")
            
            # 加载文件A
            if self.file_a_path.get():
                self.df_a = pd.read_excel(self.file_a_path.get())
                self.columns_a = list(self.df_a.columns)
                self.combo_a['values'] = self.columns_a
                if self.columns_a:
                    self.combo_a.set(self.columns_a[0])
                    
            # 加载文件B
            if self.file_b_path.get():
                self.df_b = pd.read_excel(self.file_b_path.get())
                self.columns_b = list(self.df_b.columns)
                self.combo_b['values'] = self.columns_b
                if self.columns_b:
                    self.combo_b.set(self.columns_b[0])
                    
            self.status_bar.config(text="文件加载完成")
            messagebox.showinfo("成功", "文件加载完成！")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件时出错：{str(e)}")
            self.status_bar.config(text="文件加载失败")
            
    def start_compare(self):
        """开始对比"""
        if self.df_a is None or self.df_b is None:
            messagebox.showwarning("警告", "请先加载文件！")
            return
            
        if not self.column_a.get() or not self.column_b.get():
            messagebox.showwarning("警告", "请选择要对比的列！")
            return
            
        # 在新线程中执行对比操作
        thread = threading.Thread(target=self.perform_compare)
        thread.daemon = True
        thread.start()
        
    def perform_compare(self):
        """执行对比操作"""
        try:
            self.status_bar.config(text="正在执行对比...")
            self.progress_var.set(0)
            
            # 获取选择的列
            col_a = self.column_a.get()
            col_b = self.column_b.get()
            
            # 获取列数据
            data_a = set(self.df_a[col_a].dropna().astype(str))
            data_b = set(self.df_b[col_b].dropna().astype(str))
            
            self.progress_var.set(30)
            
            # 找出差异
            only_in_a = data_a - data_b  # A中有B中没有的
            only_in_b = data_b - data_a  # B中有A中没有的
            
            self.progress_var.set(60)
            
            # 创建结果DataFrame
            result_df = self.df_a.copy()
            
            # 标记A中独有的数据（标红）
            mask_a = result_df[col_a].astype(str).isin(only_in_a)
            result_df['差异标记'] = ''
            result_df.loc[mask_a, '差异标记'] = 'A中独有'
            
            # 添加B中独有的数据
            b_only_data = self.df_b[self.df_b[col_b].astype(str).isin(only_in_b)]
            if not b_only_data.empty:
                # 创建新行，填充B的数据
                new_rows = []
                for _, row in b_only_data.iterrows():
                    new_row = pd.Series(index=result_df.columns, dtype=object)
                    new_row[col_a] = row[col_b]  # 将B的列数据放入A的列
                    new_row['差异标记'] = 'B中独有'
                    new_rows.append(new_row)
                
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    result_df = pd.concat([result_df, new_df], ignore_index=True)
            
            self.progress_var.set(90)
            
            # 保存结果
            self.result_df = result_df
            
            # 更新结果显示
            self.root.after(0, self.update_result_display, only_in_a, only_in_b)
            
            self.progress_var.set(100)
            self.status_bar.config(text="对比完成")
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"对比过程中出错：{str(e)}"))
            self.status_bar.config(text="对比失败")
            
    def update_result_display(self, only_in_a, only_in_b):
        """更新结果显示"""
        self.result_text.delete(1.0, tk.END)
        
        result_text = "对比结果：\n\n"
        result_text += f"A文件中独有数据（{len(only_in_a)}条）：\n"
        for item in sorted(only_in_a):
            result_text += f"  - {item}\n"
            
        result_text += f"\nB文件中独有数据（{len(only_in_b)}条）：\n"
        for item in sorted(only_in_b):
            result_text += f"  - {item}\n"
            
        result_text += f"\n总计差异：{len(only_in_a) + len(only_in_b)}条"
        
        self.result_text.insert(1.0, result_text)
        messagebox.showinfo("完成", "对比操作完成！")
        
    def save_result(self):
        """保存对比结果"""
        if not hasattr(self, 'result_df'):
            messagebox.showwarning("警告", "没有可保存的结果！")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存对比结果",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if file_path:
                # 使用openpyxl保存，支持样式
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    self.result_df.to_excel(writer, index=False, sheet_name='对比结果')
                    
                    # 获取工作表
                    worksheet = writer.sheets['对比结果']
                    
                    # 设置红色填充样式
                    red_fill = PatternFill(start_color='FFFF0000', 
                                         end_color='FFFF0000', 
                                         fill_type='solid')
                    
                    # 为差异标记的单元格设置红色背景
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                        if row[-1].value in ['A中独有', 'B中独有']:  # 最后一列是差异标记
                            for cell in row:
                                cell.fill = red_fill
                
                messagebox.showinfo("成功", f"结果已保存到：{file_path}")
                self.status_bar.config(text="结果保存完成")
                
        except Exception as e:
            messagebox.showerror("错误", f"保存文件时出错：{str(e)}")
            self.status_bar.config(text="保存失败")
    
    # 数据去重相关方法
    def select_dedup_file(self):
        """选择去重文件"""
        file_path = filedialog.askopenfilename(
            title="选择要去重的Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if file_path:
            self.dedup_file_path.set(file_path)
    
    def load_dedup_file(self):
        """加载去重文件"""
        try:
            if not self.dedup_file_path.get():
                messagebox.showwarning("警告", "请先选择文件！")
                return
                
            self.status_bar.config(text="正在加载文件...")
            
            self.dedup_df = pd.read_excel(self.dedup_file_path.get())
            dedup_columns = list(self.dedup_df.columns)
            self.dedup_combo['values'] = dedup_columns
            if dedup_columns:
                self.dedup_combo.set(dedup_columns[0])
                
            self.status_bar.config(text="文件加载完成")
            messagebox.showinfo("成功", "文件加载完成！")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件时出错：{str(e)}")
            self.status_bar.config(text="文件加载失败")
    
    def start_deduplication(self):
        """开始去重"""
        if not hasattr(self, 'dedup_df'):
            messagebox.showwarning("警告", "请先加载文件！")
            return
            
        if not self.dedup_column.get():
            messagebox.showwarning("警告", "请选择要去重的列！")
            return
            
        # 在新线程中执行去重操作
        thread = threading.Thread(target=self.perform_deduplication)
        thread.daemon = True
        thread.start()
    
    def perform_deduplication(self):
        """执行去重操作"""
        try:
            self.status_bar.config(text="正在执行去重...")
            
            # 获取去重列
            dedup_col = self.dedup_column.get()
            strategy = self.keep_strategy.get()
            
            # 记录原始数据量
            original_count = len(self.dedup_df)
            
            # 执行去重
            if strategy == "first":
                result_df = self.dedup_df.drop_duplicates(subset=[dedup_col], keep='first')
            else:
                result_df = self.dedup_df.drop_duplicates(subset=[dedup_col], keep='last')
            
            # 记录去重后数据量
            deduped_count = len(result_df)
            removed_count = original_count - deduped_count
            
            # 保存结果
            self.dedup_result_df = result_df
            
            # 更新结果显示
            self.root.after(0, self.update_dedup_result_display, original_count, deduped_count, removed_count)
            
            self.status_bar.config(text="去重完成")
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"去重过程中出错：{str(e)}"))
            self.status_bar.config(text="去重失败")
    
    def update_dedup_result_display(self, original_count, deduped_count, removed_count):
        """更新去重结果显示"""
        self.dedup_result_text.delete(1.0, tk.END)
        
        result_text = "去重结果：\n\n"
        result_text += f"原始数据量：{original_count}条\n"
        result_text += f"去重后数据量：{deduped_count}条\n"
        result_text += f"删除重复数据：{removed_count}条\n"
        result_text += f"去重率：{(removed_count/original_count*100):.2f}%\n"
        
        self.dedup_result_text.insert(1.0, result_text)
        messagebox.showinfo("完成", "去重操作完成！")
    
    def save_dedup_result(self):
        """保存去重结果"""
        if not hasattr(self, 'dedup_result_df'):
            messagebox.showwarning("警告", "没有可保存的结果！")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存去重结果",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if file_path:
                self.dedup_result_df.to_excel(file_path, index=False)
                messagebox.showinfo("成功", f"结果已保存到：{file_path}")
                self.status_bar.config(text="去重结果保存完成")
                
        except Exception as e:
            messagebox.showerror("错误", f"保存文件时出错：{str(e)}")
            self.status_bar.config(text="保存失败")
    
    # 格式调整相关方法
    def select_format_file(self):
        """选择格式调整文件"""
        file_path = filedialog.askopenfilename(
            title="选择要调整格式的Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if file_path:
            self.format_file_path.set(file_path)
    
    def load_format_file(self):
        """加载格式调整文件"""
        try:
            if not self.format_file_path.get():
                messagebox.showwarning("警告", "请先选择文件！")
                return
                
            self.status_bar.config(text="正在加载文件...")
            
            self.format_df = pd.read_excel(self.format_file_path.get())
            format_columns = list(self.format_df.columns)
            self.format_combo['values'] = format_columns
            if format_columns:
                self.format_combo.set(format_columns[0])
                
            self.status_bar.config(text="文件加载完成")
            messagebox.showinfo("成功", "文件加载完成！")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件时出错：{str(e)}")
            self.status_bar.config(text="文件加载失败")
    
    def apply_format(self):
        """应用格式调整"""
        if not hasattr(self, 'format_df'):
            messagebox.showwarning("警告", "请先加载文件！")
            return
            
        if not self.format_column.get():
            messagebox.showwarning("警告", "请选择目标列！")
            return
            
        # 在新线程中执行格式调整
        thread = threading.Thread(target=self.perform_format_adjustment)
        thread.daemon = True
        thread.start()
    
    def perform_format_adjustment(self):
        """执行格式调整"""
        try:
            self.status_bar.config(text="正在调整格式...")
            
            # 获取目标列和格式类型
            target_col = self.format_column.get()
            format_type = self.format_type.get()
            
            # 复制数据
            result_df = self.format_df.copy()
            
            # 根据格式类型进行调整
            if format_type == "date":
                # 日期格式调整
                try:
                    result_df[target_col] = pd.to_datetime(result_df[target_col], errors='coerce')
                    result_df[target_col] = result_df[target_col].dt.strftime('%Y-%m-%d')
                    format_result = "日期格式已统一为 YYYY-MM-DD"
                except:
                    format_result = "日期格式调整失败，请检查数据格式"
                    
            elif format_type == "number":
                # 数字格式调整
                try:
                    result_df[target_col] = pd.to_numeric(result_df[target_col], errors='coerce')
                    result_df[target_col] = result_df[target_col].round(2)  # 保留两位小数
                    format_result = "数字格式已调整，保留两位小数"
                except:
                    format_result = "数字格式调整失败，请检查数据格式"
                    
            elif format_type == "text":
                # 文本格式调整
                try:
                    result_df[target_col] = result_df[target_col].astype(str).str.strip()
                    format_result = "文本格式已调整，去除首尾空格"
                except:
                    format_result = "文本格式调整失败"
            
            # 保存结果
            self.format_result_df = result_df
            
            # 更新结果显示
            self.root.after(0, self.update_format_result_display, format_result)
            
            self.status_bar.config(text="格式调整完成")
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"格式调整过程中出错：{str(e)}"))
            self.status_bar.config(text="格式调整失败")
    
    def update_format_result_display(self, format_result):
        """更新格式调整结果显示"""
        self.format_result_text.delete(1.0, tk.END)
        
        result_text = "格式调整结果：\n\n"
        result_text += f"调整类型：{self.format_type.get()}\n"
        result_text += f"目标列：{self.format_column.get()}\n"
        result_text += f"调整结果：{format_result}\n"
        result_text += f"处理记录数：{len(self.format_result_df)}条\n"
        
        self.format_result_text.insert(1.0, result_text)
        messagebox.showinfo("完成", "格式调整完成！")
    
    def save_format_result(self):
        """保存格式调整结果"""
        if not hasattr(self, 'format_result_df'):
            messagebox.showwarning("警告", "没有可保存的结果！")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存格式调整结果",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if file_path:
                self.format_result_df.to_excel(file_path, index=False)
                messagebox.showinfo("成功", f"结果已保存到：{file_path}")
                self.status_bar.config(text="格式调整结果保存完成")
                
        except Exception as e:
            messagebox.showerror("错误", f"保存文件时出错：{str(e)}")
            self.status_bar.config(text="保存失败")
            
    def run(self):
        """运行应用程序"""
        self.root.mainloop()

if __name__ == "__main__":
    app = ExcelTool()
    app.run()